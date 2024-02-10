from openpyxl import Workbook
from openpyxl.styles import Font

from events.services import (
    get_event_participants,
    get_event_teams,
    get_participant_solution,
    get_team_participants,
    get_team_participants_email_string,
    get_team_participants_fio_string,
    get_team_participants_phone_number_string,
    get_team_solution,
)


def export_event_to_excel(event) -> None:
    """Export participants data to excel table"""
    if event.type == 'Индивидуальное':
        create_workbook_for_individual_event(event=event)
    else:
        create_workbook_for_team_event(event=event)


def create_workbook_for_individual_event(event):
    wb = Workbook()
    ws = wb.active

    ws.title = f'{event.slug}'

    ws['A1'] = 'ФИО ученика'
    ws['B1'] = 'Школа ученика'
    ws['C1'] = 'Почта ученика'
    ws['D1'] = 'Телефон ученика'
    ws['E1'] = 'ФИО руководителя'
    ws['F1'] = 'Почта руководителя'
    ws['G1'] = 'Телефон руководителя'
    ws['H1'] = 'Ссылка на приложенные файлы'
    ws['I1'] = 'Тема проекта'
    ws['J1'] = 'Предмет'
    ws['K1'] = 'Год обучения ученика'

    for cell in ws[1]:
        cell.font = Font(bold=True)

    letter_string = 'ABCDEFGHIJK'
    participant_index = 1
    participants = get_event_participants(event=event)

    for participant in participants:
        participant_index = participant_index + 1
        for index in range(11):
            cell = letter_string[index] + str(participant_index)
            solution = get_participant_solution(event=event, participant=participant)
            if letter_string[index] == 'A':
                ws[cell] = str(
                    participant.user.full_name,
                )
            if letter_string[index] == 'B':
                ws[cell] = str(
                    participant.user.profile.school,
                )
            if letter_string[index] == 'C':
                ws[cell] = str(
                    participant.user.email,
                )
            if letter_string[index] == 'D':
                ws[cell] = str(
                    participant.user.profile.phone_number,
                )
            if letter_string[index] == 'E':
                ws[cell] = str(
                    participant.supervisor_fio,
                )
            if letter_string[index] == 'F':
                ws[cell] = str(
                    participant.supervisor_email,
                )
            if letter_string[index] == 'G':
                ws[cell] = str(
                    participant.supervisor_phone_number,
                )
            if letter_string[index] == 'H':
                ws[cell] = str(
                    solution.url if solution else '',
                )
            if letter_string[index] == 'I':
                ws[cell] = str(
                    solution.topic if solution else '',
                )
            if letter_string[index] == 'J':
                ws[cell] = str(
                    solution.subject if solution else '',
                )
            if letter_string[index] == 'K':
                ws[cell] = str(
                    participant.user.profile.year_of_study,
                )

    wb.save(f'media/event_{event.id}.xlsx')


def create_workbook_for_team_event(event):
    wb = Workbook()
    ws = wb.active

    ws.title = f'{event.slug}'

    ws['A1'] = 'Название команды'
    ws['B1'] = 'Название класса'
    ws['C1'] = 'Школа учеников'
    ws['D1'] = 'ФИО учеников'
    ws['E1'] = 'Почты учеников'
    ws['F1'] = 'Телефоны учеников'
    ws['G1'] = 'ФИО руководителя'
    ws['H1'] = 'Почта руководителя'
    ws['I1'] = 'Телефон руководителя'
    ws['J1'] = 'Ссылка на приложенные файлы'
    ws['K1'] = 'Тема проекта'
    ws['L1'] = 'Предмет'

    for cell in ws[1]:
        cell.font = Font(bold=True)

    letter_string = 'ABCDEFGHIJKL'
    team_index = 1
    teams = get_event_teams(event=event)

    for team in teams:
        team_index = team_index + 1
        for index in range(12):
            solution = get_team_solution(event=event, team=team)
            cell = letter_string[index] + str(team_index)
            if letter_string[index] == 'A':
                ws[cell] = str(
                    team.name,
                )
            if letter_string[index] == 'B':
                ws[cell] = str(
                    team.school_class,
                )
            if letter_string[index] == 'C':
                if team.supervisor:
                    ws[cell] = str(
                        team.supervisor.profile.school,
                    )
                elif get_team_participants(team=team):
                    try:
                        ws[cell] = str(
                            get_team_participants(
                                team=team,
                            ).first().user.profile.school,
                        )
                    except:  # noqa E722
                        ws[cell] = ''
                else:
                    ws[cell] = ''
            if letter_string[index] == 'D':
                ws[cell] = str(
                    get_team_participants_fio_string(team=team),
                )
            if letter_string[index] == 'E':
                ws[cell] = str(
                    get_team_participants_email_string(team=team),
                )
            if letter_string[index] == 'F':
                ws[cell] = str(
                    get_team_participants_phone_number_string(team=team),
                )
            if letter_string[index] == 'G':
                ws[cell] = str(
                    team.supervisor_fio,
                )
            if letter_string[index] == 'H':
                ws[cell] = str(
                    team.supervisor_email,
                )
            if letter_string[index] == 'I':
                ws[cell] = str(
                    team.supervisor_phone_number,
                )
            if letter_string[index] == 'J':
                ws[cell] = str(
                    solution.url if solution else '',
                )
            if letter_string[index] == 'K':
                ws[cell] = str(
                    solution.topic if solution else '',
                )
            if letter_string[index] == 'L':
                ws[cell] = str(
                    solution.subject if solution else '',
                )

    wb.save(f'media/event_{event.id}.xlsx')
